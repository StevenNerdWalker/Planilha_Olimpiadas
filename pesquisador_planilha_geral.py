import openpyxl as xls
import webbrowser
import time
import datetime

'''links for modules documentation:
https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
https://docs.python.org/3/library/webbrowser.html'''

class Event:
    def __init__(self, name:str, date:datetime.date, description:str):
        self.name = name
        self.date = date
        self.description = description
    def get_name(self):
        return self.name
    def get_date(self):
        return self.date
    def get_description(self):
        return self.description
    
def get_events(filepath:str, sheet:str, col: int, line: int):
    '''Takes in the path of the file, the sheet containing the events, the column with the events names,
    the line with the first event.
    
    The columns MUST be in this order: event name, date, description.
    If not, the code WILL NOT work properly.'''

    wb = xls.load_workbook(filename=filepath)
    ws = wb[sheet]

    events_list = []
    while True:
        name = ws.cell(column=col, row=line).value
        date = ws.cell(column=col+1, row=line).value
        description = ws.cell(column=col+2, row=line).value

        if name == None: break
        if line > 100: break

        event = Event(name, date, description)

        events_list.append(event)
        line += 1

    return events_list

def search(filepath:str, sheet:str, col: int, line: int, search_term: str):
    '''Takes in the path of the file, the sheet containing the events, the column with the events names,
    the line with the first event, and a search term to specify the searches.
    
    The columns MUST be in this order: event name, date, description.
    If not, the code may not work properly.'''

    event_list = get_events(filepath, sheet, col, line)

    names_set = set()
    for event in event_list:
        event_name = event.get_name()
        names_set.add(event_name)

    for name in names_set:
        webbrowser.open(url=f'https://duckduckgo.com/?t=ffab&q={name}+{search_term}&atb=v388-1&ia=web')
        time.sleep(3)

def sort_events(event_list:list):
    unsorted_event_list = event_list.copy()
    sorted_event_list = []

    for i in range(0, len(event_list)):
        earliest_event = unsorted_event_list[0]
        for j in range(0, len(unsorted_event_list)):
            variable_event = unsorted_event_list[j]
            if variable_event.get_date() < earliest_event.get_date():
                earliest_event = variable_event
        sorted_event_list.append(earliest_event)
        unsorted_event_list.remove(earliest_event)

    return sorted_event_list

    
def order(filepath:str, sheet:str, col:int, line:int):
    unsorted_event_list = get_events(filepath, sheet, col, line)
    
    sorted_event_list = sort_events(unsorted_event_list)

    wb = xls.load_workbook(filename=filepath)
    ws = wb[sheet]

    for i in range(0,len(sorted_event_list)):
        event = sorted_event_list[i]

        name = event.get_name()
        date = event.get_date()
        description = event.get_description()

        curr_line = line + i

        namecell = ws.cell(column=col, row=curr_line)
        namecell.value = name
        datecell = ws.cell(column=col+1, row=curr_line)
        datecell.value = date
        descriptioncell = ws.cell(column=col+2, row=curr_line)
        descriptioncell.value = description
    wb.save(filename=filepath)
